import openpyxl
import os
import sys
from copy import copy
from datetime import date, datetime
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font

import config
import util


def combiner(itemkey, methodname, *a, **k):
    def keyextractor(container):
        item = container[itemkey]
        method = getattr(item, methodname)
        return method

    return keyextractor


def openxyl_row_copy(ws, row, strike=False):
    copied_row = []
    for cell in row:
        new_cell = Cell(ws, value=cell.value)
        new_cell.style = copy(cell.style)
        new_cell.font = new_cell.font + Font(strike=strike)
        copied_row.append(new_cell)
    return copied_row


def _load_from_result_file(result_file_name):
    free_list = []
    lately_list = []
    top_ranked_list = []
    group_by_book_code = {}
    over_20k_favorite_count_set = set()
    over_15k_favorite_count_set = set()
    over_10k_favorite_count_set = set()
    xlsx_obj = openpyxl.load_workbook(result_file_name)
    for ws in xlsx_obj.worksheets:
        in_free_list = in_lately_list = False
        for row in ws.rows:
            first_cell_value = row[0].value

            if first_cell_value == "무료 순위":
                in_free_list = True
                continue
            elif first_cell_value == "신규 순위":
                in_lately_list = True
                continue

            try:
                ranking = int(first_cell_value)
            except:
                continue

            if in_free_list:
                free_list.append(row)
            elif in_lately_list:
                lately_list.append(row)
            if ranking == 1:
                top_ranked_list.append(row)

            book_code = int(row[4].value)
            book_total_favorite_count = int(row[13].value)
            if book_total_favorite_count >= 20000:
                over_20k_favorite_count_set.add(book_code)
            elif book_total_favorite_count >= 15000:
                over_15k_favorite_count_set.add(book_code)
            elif book_total_favorite_count >= 10000:
                over_10k_favorite_count_set.add(book_code)

            if book_code not in group_by_book_code:
                group_by_book_code[book_code] = []

            group_by_book_code[book_code].append(row)

    for book_code in group_by_book_code:
        book_rows = group_by_book_code[book_code]
        group_by_book_code[book_code] = sorted(book_rows, key=combiner(3, 'value'))

    top_ranked_list = sorted(top_ranked_list, key=combiner(3, 'value'))

    return {
        "free_list": free_list,
        "lately_list": lately_list,
        "top_ranked_list": top_ranked_list,
        "over_20k_favorite_count_set": over_20k_favorite_count_set,
        "over_15k_favorite_count_set": over_15k_favorite_count_set,
        "over_10k_favorite_count_set": over_10k_favorite_count_set,
        "group_by_book_code": group_by_book_code,
    }


def create_summary_sheet(xlsx_obj, all_rows, book_code, ws_name):
    if book_code not in all_rows["group_by_book_code"]:
        print(f"{book_code} not found.")
        return
    sheet = xlsx_obj.create_sheet()
    headers = config.DATA_HEADERS
    sheet.append(headers)
    for book_row in all_rows["group_by_book_code"][book_code]:
        sheet.append(openxyl_row_copy(sheet, book_row))
    max_row = sheet.max_row
    sheet.cell(max_row + 1, 10, f'=AVERAGEIF(H2:H{max_row},"<20",J2:J{max_row})')
    sheet.cell(max_row + 1, 12, f'=AVERAGEIF(H2:H{max_row},"<20",L2:L{max_row})')
    sheet.title = f"{ws_name}_{book_code}"


def create_always_include_summary_sheet(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows):
    for book_code in config.ALWAYS_ANALYZE_BOOK_CODE_LIST:
        create_summary_sheet(xlsx_obj_20k, all_rows, book_code, "작품 분석")
        create_summary_sheet(xlsx_obj_15k, all_rows, book_code, "작품 분석")
        create_summary_sheet(xlsx_obj_10k, all_rows, book_code, "작품 분석")


def create_summary_sheets(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows):
    create_always_include_summary_sheet(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows)
    for book_code in all_rows["over_20k_favorite_count_set"]:
        create_summary_sheet(xlsx_obj_20k, all_rows, book_code, "2만작 분석")
    for book_code in all_rows["over_15k_favorite_count_set"]:
        create_summary_sheet(xlsx_obj_15k, all_rows, book_code, "1.5만작 분석")
    for book_code in all_rows["over_10k_favorite_count_set"]:
        create_summary_sheet(xlsx_obj_10k, all_rows, book_code, "1만작 분석")


def create_calendar_sheet(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows):
    pass


def create_monthly_sheet(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows):
    top_ranked_book_code_set = set()
    for top_ranked in all_rows["top_ranked_list"]:
        target_date = datetime.strptime(top_ranked[3].value, "%y%m%d")
        ws_name = f"{target_date.year}년{target_date.month}월_1위"
        if ws_name in xlsx_obj_20k:
            ws_20k = xlsx_obj_20k[ws_name]
        else:
            ws_20k = xlsx_obj_20k.create_sheet(title=ws_name)
        if ws_name in xlsx_obj_15k:
            ws_15k = xlsx_obj_15k[ws_name]
        else:
            ws_15k = xlsx_obj_15k.create_sheet(title=ws_name)
        if ws_name in xlsx_obj_10k:
            ws_10k = xlsx_obj_10k[ws_name]
        else:
            ws_10k = xlsx_obj_10k.create_sheet(title=ws_name)

        book_code = top_ranked[4].value
        strike = book_code in top_ranked_book_code_set
        top_ranked_book_code_set.add(book_code)
        ws_20k.append(openxyl_row_copy(ws_20k, top_ranked, strike))
        ws_15k.append(openxyl_row_copy(ws_15k, top_ranked, strike))
        ws_10k.append(openxyl_row_copy(ws_10k, top_ranked, strike))


def main(result_file_name, analyzed_file_name):
    all_rows = _load_from_result_file(result_file_name)

    execution_datetime_str = str(datetime.now())[0:19]

    filename, file_extension = os.path.splitext(analyzed_file_name)
    analyzed_file_name_now_20k = f'{filename}_2만작_{execution_datetime_str.replace(":", "-")}{file_extension}'
    analyzed_file_name_now_15k = f'{filename}_1.5만작_{execution_datetime_str.replace(":", "-")}{file_extension}'
    analyzed_file_name_now_10k = f'{filename}_1만작_{execution_datetime_str.replace(":", "-")}{file_extension}'

    xlsx_obj_20k = openpyxl.Workbook()
    xlsx_obj_15k = openpyxl.Workbook()
    xlsx_obj_10k = openpyxl.Workbook()

    util.add_styles(xlsx_obj_20k)
    util.add_styles(xlsx_obj_15k)
    util.add_styles(xlsx_obj_10k)

    for ws in xlsx_obj_20k.worksheets:
        xlsx_obj_20k.remove(ws)
    for ws in xlsx_obj_15k.worksheets:
        xlsx_obj_15k.remove(ws)
    for ws in xlsx_obj_10k.worksheets:
        xlsx_obj_10k.remove(ws)

    create_summary_sheets(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows)
    create_calendar_sheet(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows)
    create_monthly_sheet(xlsx_obj_20k, xlsx_obj_15k, xlsx_obj_10k, all_rows)

    xlsx_obj_20k.save(analyzed_file_name_now_20k)
    xlsx_obj_15k.save(analyzed_file_name_now_15k)
    xlsx_obj_10k.save(analyzed_file_name_now_10k)


if __name__ == "__main__":
    today = date.today()
    result_xlsx_file_name = sys.argv[1] if len(sys.argv) >= 2 else config.joara.DEFAULT_RESULT_XLSX_FILE_NAME
    analyzed_xlsx_file_name = sys.argv[2] if len(sys.argv) >= 3 else config.joara.DEFAULT_ANALYZED_XLSX_FILE_NAME
    main(result_xlsx_file_name, analyzed_xlsx_file_name)
