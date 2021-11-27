import os
import sys
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.styles.colors import BLUE, RED

import config
import scraper


def styled_cells(data, ws, style):
    for c in data:
        c = Cell(ws, value=c)
        c.style = style
        yield c


def main(date_str, file_name):
    sheet_rows = []
    execution_datetime_str = str(datetime.now())[0:19]
    line = f"조아라 투베 자동 수집기 - 실행일시 : {execution_datetime_str}"
    sheet_rows.append([line])
    print(line)
    line = f"조회 날짜 : {date_str}"
    sheet_rows.append([line])
    print(line)

    date_obj = {
        'cur_year': date_str[0:2],
        'cur_month': date_str[2:4],
        'cur_day': date_str[4:6],
    }

    if os.path.isfile(file_name):
        xlsx_obj = openpyxl.load_workbook(file_name)
    else:
        xlsx_obj = openpyxl.Workbook()

    sheet = xlsx_obj.create_sheet(title=f'{date_str} {execution_datetime_str.replace(":", "-")}')

    sheet_rows = sheet_rows + scraper.today_best.get_free_list(date_obj)
    sheet_rows = sheet_rows + scraper.today_best.get_lately_list(date_obj)

    if "over_5k_style" not in xlsx_obj.style_names:
        over_5k_favorite_color = "008000"
        over_5k_favorite_bg_color = "B2B2B2"

        over_5k_style = NamedStyle(name="over_5k_style")
        over_5k_style.font = Font(color=over_5k_favorite_color)
        over_5k_style.fill = PatternFill(start_color=over_5k_favorite_bg_color,
                                          end_color=over_5k_favorite_bg_color,
                                          fill_type="solid")

        xlsx_obj.add_named_style(over_10k_style)

    if "over_10k_style" not in xlsx_obj.style_names:
        over_10k_favorite_color = BLUE
        over_10k_favorite_bg_color = "F0E8DD"

        over_10k_style = NamedStyle(name="over_10k_style")
        over_10k_style.font = Font(color=over_10k_favorite_color)
        over_10k_style.fill = PatternFill(start_color=over_10k_favorite_bg_color,
                                          end_color=over_10k_favorite_bg_color,
                                          fill_type="solid")

        xlsx_obj.add_named_style(over_10k_style)

    if "over_20k_style" not in xlsx_obj.style_names:
        over_20k_favorite_color = RED
        over_20k_favorite_bg_color = "D1F5EC"

        over_20k_style = NamedStyle(name="over_20k_style")
        over_20k_style.font = Font(color=over_20k_favorite_color)
        over_20k_style.fill = PatternFill(start_color=over_20k_favorite_bg_color,
                                          end_color=over_20k_favorite_bg_color,
                                          fill_type="solid")

        xlsx_obj.add_named_style(over_20k_style)

    for row in sheet_rows:
        row_style = None
        if len(row) > 14:
            try:
                book_total_favorite_count = row[13]
                if book_total_favorite_count >= 20000:
                    row_style = "over_20k_style"
                elif book_total_favorite_count >= 10000:
                    row_style = "over_10k_style"
                elif book_total_favorite_count >= 5000:
                    row_style = "over_5k_style"

                if row_style is not None:
                    row = styled_cells(row, sheet, row_style)
            except:
                pass
        sheet.append(row)

    for row_index in list(range(0, sheet.max_row)):
        sheet.row_dimensions[row_index].height = 15

    sheet["J105"] = "=SUM(J5:J104)"
    sheet["L105"] = "=SUM(L5:L104)"

    xlsx_obj.save(file_name)


if __name__ == "__main__":
    today = date.today()
    start_date_string = sys.argv[1] if len(sys.argv) >= 2 else f"{today.year%100:02d}{today.month:02d}{today.day:02d}"
    end_date_string = sys.argv[2] if len(sys.argv) >= 3 else start_date_string
    xlsx_file_name = sys.argv[3] if len(sys.argv) >= 4 else config.joara.DEFAULT_RESULT_XLSX_FILE_NAME
    start_date = datetime.strptime(start_date_string, "%y%m%d")
    end_date = datetime.strptime(end_date_string, "%y%m%d")
    one_day = timedelta(days=1)
    while end_date >= start_date:
        main(start_date.strftime("%y%m%d"), xlsx_file_name)
        start_date += one_day
