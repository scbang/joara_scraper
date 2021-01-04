import os
import sys
from datetime import *
from typing import Dict, List

import openpyxl
from dateutil.tz import *

import config
import scraper
from config import make_url
from data_object.book import RidibooksBook

datetime.now(tzlocal()).isoformat()


def append_line_to_work_sheet(
        work_sheet_rows: List[List[str]],
        lines: List[List[str]],
) -> int:
    n_line = 0
    for line in lines:
        work_sheet_rows.append(line)
        n_line += 1
    return n_line


def append_book_details_to_work_sheet(
        work_sheet_rows: List[List[str]],
        book_item_list: List[Dict],
        book_details: Dict[str, Dict[str, RidibooksBook]],
        merge_infos: List[Dict],
        section_name: str,
        given_row_index: int,
) -> int:
    current_row_index = given_row_index

    headers = []
    for header in config.ridi.BOOK_DATA_HEADERS:
        if section_name not in header[3]:
            headers.append(header)
    current_row_index += append_line_to_work_sheet(work_sheet_rows, [[header[0] for header in headers]])

    for i, book_item in enumerate(book_item_list):
        merge_start_row_index = current_row_index
        book_detail = book_details[section_name][book_item['b_id']]
        found_in_today_recommendation = "O" if book_item['b_id'] in book_details['today_recommendation_list'] else "X"
        found_in_today_new = "O" if book_item['b_id'] in book_details['today_new_list'] else "X"
        for j, author in enumerate(book_detail.authors):
            sheet_row = []
            recent_published_book = author.recent_published_book
            for header in headers:
                sheet_row.append(eval(header[1]) if j == 0 or not header[2] else "")
            current_row_index += append_line_to_work_sheet(work_sheet_rows, [sheet_row])

        if len(book_detail.authors) > 1:
            for k, header in enumerate(headers):
                if not header[2]:  # do merge?
                    continue
                merge_infos.append({
                    "start_row": merge_start_row_index + 1,
                    "start_column": k + 1,
                    "end_row": current_row_index,
                    "end_column": k + 1,
                })

    return current_row_index - given_row_index


def get_excel_rows(
        work_sheet_rows: List[List[str]],
        book_details: Dict,
        today_recommendation_list: List,
        today_new_list: List,
        event_books: List,
        publishers: Dict,
) -> List[Dict]:
    merge_infos = []

    current_row_index = len(work_sheet_rows)

    if today_recommendation_list:
        current_row_index += append_line_to_work_sheet(work_sheet_rows, [
            [],
            [f"+++ [오늘, 리디의 발견] 수집 결과, {len(today_recommendation_list)}개의 책 페이지 발견"],
        ])
        current_row_index += append_book_details_to_work_sheet(
            work_sheet_rows,
            today_recommendation_list,
            book_details,
            merge_infos,
            'today_recommendation_list',
            current_row_index,
        )

    if today_new_list:
        current_row_index += append_line_to_work_sheet(work_sheet_rows, [
            [],
            [f"+++ [오늘의 신간] 수집 결과, {len(today_new_list)}개의 책 페이지 발견"],
        ])
        current_row_index += append_book_details_to_work_sheet(
            work_sheet_rows,
            today_new_list,
            book_details,
            merge_infos,
            'today_new_list',
            current_row_index,
        )

    if event_books:
        current_row_index += append_line_to_work_sheet(work_sheet_rows, [
            [],
            [f"+++ 최상단 배너 영역 발견 수집 결과. {len(event_books)}개 이벤트 노출 중"],
        ])
        for i, item in enumerate(event_books):
            top_banner_event = item["event_info"]
            event_book_list = item["book_items"]

            current_row_index += append_line_to_work_sheet(work_sheet_rows, [
                [],
                [f"--- {i+1}번째 이벤트, [{top_banner_event['title']}]"
                 f", 링크 = {make_url(top_banner_event['url'])}"
                 f", {len(event_book_list)}개의 책 페이지 발견"],
            ])

            current_row_index += append_book_details_to_work_sheet(
                work_sheet_rows,
                event_book_list,
                book_details,
                merge_infos,
                'event_books',
                current_row_index,
            )

    publisher_lines = []
    for publisher_name, publisher_obj in publishers.items():
        publisher_line = []
        for header in config.ridi.PUBLISHER_DETAIL_HEADERS:
            publisher_line.append(eval(header[1]))
        publisher_lines.append(publisher_line)

    current_row_index += append_line_to_work_sheet(work_sheet_rows, [
        [],
        [f"출판사 {len(publishers)}개 수집됨. {list(publishers)}"],
        [header[0] for header in config.ridi.PUBLISHER_DETAIL_HEADERS],
    ] + publisher_lines)
    return merge_infos


def main(_id: str, password: str, file_name: str):
    xlsx_obj = openpyxl.load_workbook(file_name) if os.path.isfile(file_name) else openpyxl.Workbook()
    execution_datetime_str = datetime.now(tzlocal()).isoformat()
    scrape_data = scraper.scrape_romance_home(_id, password)
    sheet_rows = [[f"리디북스 로맨스 e북 메인 페이지 데이터 수집기 - 실행일시 : {execution_datetime_str}"]]
    merge_infos = get_excel_rows(sheet_rows, **scrape_data)
    sheet = xlsx_obj.create_sheet(
        title=f'{(execution_datetime_str[0:19] + execution_datetime_str[25:]).replace(":", "-")}'
    )
    for row in sheet_rows:
        sheet.append(row)
    for merge_info in merge_infos:
        sheet.merge_cells(**merge_info)
    xlsx_obj.save(file_name)


def test(_id: str, password: str):
    scraper.test(_id, password)


if __name__ == "__main__":
    account_id = sys.argv[1]
    account_password = sys.argv[2]
    xlsx_file_name = sys.argv[3] if len(sys.argv) >= 4 else config.ridi.DEFAULT_RESULT_XLSX_FILE_NAME

    main(account_id, account_password, xlsx_file_name)
    # test(account_id, account_password)
