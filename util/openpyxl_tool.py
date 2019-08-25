from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.styles.colors import BLUE, RED


def add_styles(xlsx_obj):
    if "over_10k_style" not in xlsx_obj.style_names:
        over_10k_favorite_color = BLUE
        over_10k_favorite_bg_color = "F0E8DD"

        over_10k_style = NamedStyle(name="over_10k_style")
        over_10k_style.font = Font(color=over_10k_favorite_color)
        over_10k_style.fill = PatternFill(start_color=over_10k_favorite_bg_color,
                                          end_color=over_10k_favorite_bg_color,
                                          fill_type="solid")

        xlsx_obj.add_named_style(over_10k_style)
    if "over_20k_style" not in xlsx_obj.style_names:
        over_20k_favorite_color = RED
        over_20k_favorite_bg_color = "D1F5EC"

        over_20k_style = NamedStyle(name="over_20k_style")
        over_20k_style.font = Font(color=over_20k_favorite_color)
        over_20k_style.fill = PatternFill(start_color=over_20k_favorite_bg_color,
                                          end_color=over_20k_favorite_bg_color,
                                          fill_type="solid")

        xlsx_obj.add_named_style(over_20k_style)
